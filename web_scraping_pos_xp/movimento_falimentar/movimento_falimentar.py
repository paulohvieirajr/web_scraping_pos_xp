import os
import time
import base64
import gzip
import re
from typing import List
from openpyxl import Workbook
from unicodedata import normalize
from urllib.parse import unquote
from playwright.sync_api import sync_playwright, Page, Locator, Browser, BrowserContext, Keyboard, TimeoutError as PlaywrightTimeoutError
from movimento_falimentar.artigo import Artigo

class MovimentoFalimentar():

    def __init__(self):
        self.url_base = "https://valor.globo.com/"

        self.LISTA_MOVIMENTOS_VALIDOS:str = 'div.paywall div[data-block-type]'

        self.INPUT_CONSULTA:str = '#top_search_text_input'
        self.BOTAO_MOVIMENTO_FALIMENTAR:str = "div .widget--info__text-container > a"
        self.BOTAO_MOVIMENTO_FALIMENTAR_VEJA_MAIS:str = "div .pagination > a"
        self.BOTAO_COOKIES_ACEITAR:str = "button .cookie-banner-lgpd_accept-button"
        self.BOTAO_PAGINACAO:str = "div .pagination > a"
        
        self.BOTAO_LOGIN:str = '.header__login__signedin.login-signedout-spot'
        self.INPUT_LOGIN_SENHA:str = '.Input__StyledInput-sc-1cbn5m9-3.NFGIy'
        self.BOTAO_CONTINUAR_ENTRAR:str = '.Button__BaseButton-sc-1dl9u2w-0.Button__PrimaryButton-sc-1dl9u2w-1.jMCuhN.dyJfsA'
        self.BOTAO_FECHAR_AVISO:str = '.close[onclick="closeWindow()"]'
        
        self.BOTAO_SAIR:str = '.login-profile .login-profile__menu .login-profile__menu__item a'
        self.BOTAO_SESSAO_SIMULTANEA:str = '#device-desktop'
        self.AVISO_SESSAO_SIMULTANEA:str = '.wall-concurrence.theme-font-primary'

    def execute(self) -> bool:
        result = False
        playwright = sync_playwright().start()

        try:
            print('Carregando playwright...')
            navegador:Browser = playwright.firefox.launch(headless=False)
            contexto:BrowserContext = navegador.new_context(no_viewport=True)
            pagina:Page = contexto.new_page()  

            print('Realizando Login...')
            resultLogin = self.realizar_login(pagina=pagina)
            if not resultLogin:
                print('Problemas no login. Tente novamnete mais tarde.')
                return result

            time.sleep(2)
            print('Login realizado...')

            self.go_to_pagina(pagina, self.url_base, 20000)
            time.sleep(2)

            print('Pesquisando movimento falimentar...')
            pagina.locator(self.INPUT_CONSULTA).fill('movimento falimentar')
            pagina.keyboard.press('Enter')
            time.sleep(5)

            try:
                sessao_simultanea:str = pagina.locator(self.AVISO_SESSAO_SIMULTANEA).inner_text(timeout=3000)
                
                if 'Detectamos que voc' in sessao_simultanea:
                    print('Multiplas sessoes detectadas...')
                    pagina.locator(self.BOTAO_SESSAO_SIMULTANEA).nth(0).click()
            
            except PlaywrightTimeoutError as e:
                pass

            try:
                pagina.wait_for_url(url='https://valor.globo.com/busca**', wait_until="load", timeout=30000)
            except :
                pass

            time.sleep(60)
            total_processado = 0
            sem_erros = True

            try:
                pagina.locator(self.BOTAO_PAGINACAO).click(timeout=3000)
            except Exception as e:
                print(f'Botao de paginacao nao localiado. Erro: {str(e)}')
                pass

            while (sem_erros):

                page_count:int = pagina.locator(self.BOTAO_MOVIMENTO_FALIMENTAR).count()

                print(f'Detectado {page_count - total_processado} paginas para processar...')

                lista_url:list[str] = []
                for i_page in range(page_count):
                    lista_url.append(pagina.locator(self.BOTAO_MOVIMENTO_FALIMENTAR).nth(i_page + total_processado).get_attribute('href'))

                wb = Workbook()

                ws0, ws0_row_index = self.__create_sheet(wb, "Falencia Requerida")
                ws1, ws1_row_index = self.__create_sheet(wb, "Falencia Decretada")
                ws2, ws2_row_index = self.__create_sheet(wb, "Processos Falencia Extintos")
                ws3, ws3_row_index = self.__create_sheet(wb, "Reformas de Sentenca de Falencia")
                ws4, ws4_row_index = self.__create_sheet(wb, "RJ Deferida")
                ws5, ws5_row_index = self.__create_sheet(wb, "RE Deferida")
                ws6, ws6_row_index = self.__create_sheet(wb, "Homologacao de Desistencia RJ")
                ws7, ws7_row_index = self.__create_sheet(wb, "Cumprimento de RJ")
                ws8, ws8_row_index = self.__create_sheet(wb, "RJ Concedidas")
                ws9, ws9_row_index = self.__create_sheet(wb, "RJ Requerida")
                ws10, ws10_row_index = self.__create_sheet(wb, "RJ Indeferida")
                ws11, ws11_row_index = self.__create_sheet(wb, "RE Requerida")
                ws12, ws12_row_index = self.__create_sheet(wb, "RE Concedidas")

                contador:int = 1
                for link in lista_url:
                    try:
                        print(f'Processando pagina {link} ...')
                        if contador == 30:
                            pagina.close()
                            contexto.close()       
                            navegador.close()
                            playwright.stop()        
                            time.sleep(2)    
                            playwright = sync_playwright().start()
                            navegador:Browser = playwright.firefox.launch(headless=False)
                            contexto:BrowserContext = navegador.new_context(no_viewport=True)
                            pagina:Page = contexto.new_page()
                            self.realizar_login(pagina=pagina)
                            contador = 1
                            time.sleep(10)                        
                            
                        time.sleep(30)
                        pagina.goto(url=unquote(f'https://{link}'), wait_until='domcontentloaded', timeout=20000)
                        time.sleep(25)

                        self.url_consulta = pagina.url

                        lista_movimentos_validos:List = pagina.locator(self.LISTA_MOVIMENTOS_VALIDOS).all()

                        data_block_type:str = ''

                        quantidade_itens:int = pagina.locator(self.LISTA_MOVIMENTOS_VALIDOS).count()

                        artigos:List[Artigo] = []
                        artigo:Artigo = Artigo()

                        consulta_movimento:list = unquote(link).split('noticia')
                        data_consulta:str = consulta_movimento[1]
                        data_consulta = data_consulta[1:11]

                        for index, i in enumerate(lista_movimentos_validos):
                            data_block_type = i.get_attribute('data-block-type')
                            pular = False
                            if data_block_type == 'raw':
                               
                                try:
                                    texto_movimento_raw:str = i.locator(".content-intertitle").inner_text()
                                    if artigo.titulo is not None:
                                        artigos.append(artigo)
                                        artigo = Artigo()

                                    artigo.titulo = texto_movimento_raw
                                except:
                                    pular = True
                                if pular:
                                    continue
                            elif data_block_type == 'unstyled':
                                try:
                                    texto_movimento_unstyled:str = i.locator("p").inner_text()
                                    artigo.paragrafos.append(texto_movimento_unstyled)
                                except:
                                    pular = True
                                if pular:
                                    continue

                            artigos.append(artigo)

                        lista_coluna1:List = []
                        lista_coluna2:List = []
                        lista_coluna3:List = []

                        for i in artigos:
                            for paragrafo in i.paragrafos:
                                padrao_cnpj = r'CNPJ: (.*?) '
                                match_cnpj = re.search(padrao_cnpj, paragrafo)
                                cnpj_lista:list = match_cnpj.group(1) if match_cnpj else None
                                campos_cnpj = f'CNPJ: {cnpj_lista} '
                                campos_cnpj = campos_cnpj.split(":")

                                paragrafo_separado:list = paragrafo.split("-")

                                lista_coluna1.append(campos_cnpj[0])
                                lista_coluna2.append(campos_cnpj[1])
                                lista_coluna3.append(i.titulo)

                                if self.__remover_acentos_(i.titulo).lower() == 'falencias requeridas':
                                    ws0.cell(row=ws0_row_index, column=1, value=data_consulta)
                                    ws0.cell(row=ws0_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'falencias decretadas':
                                    ws1.cell(row=ws1_row_index, column=1, value=data_consulta)
                                    ws1.cell(row=ws1_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'processos de falencia extintos':
                                    ws2.cell(row=ws2_row_index, column=1, value=data_consulta)
                                    ws2.cell(row=ws2_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'reformas de sentenca de falencia':
                                    ws3.cell(row=ws3_row_index, column=1, value=data_consulta)
                                    ws3.cell(row=ws3_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao judicial deferida':
                                    ws4.cell(row=ws4_row_index, column=1, value=data_consulta)
                                    ws4.cell(row=ws4_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao extrajudicial deferida':
                                    ws5.cell(row=ws5_row_index, column=1, value=data_consulta)
                                    ws5.cell(row=ws5_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'homologacao de desistencia de recuperacao judicial':
                                    ws6.cell(row=ws6_row_index, column=1, value=data_consulta)
                                    ws6.cell(row=ws6_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'cumprimento de recuperacao judicial':
                                    ws7.cell(row=ws7_row_index, column=1, value=data_consulta)
                                    ws7.cell(row=ws7_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes judiciais concedidas':
                                    ws8.cell(row=ws8_row_index, column=1, value=data_consulta)
                                    ws8.cell(row=ws8_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao judicial requerida':
                                    ws9.cell(row=ws9_row_index, column=1, value=data_consulta)
                                    ws9.cell(row=ws9_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes judiciais indeferidas':
                                    ws10.cell(row=ws10_row_index, column=1, value=data_consulta)
                                    ws10.cell(row=ws10_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao extrajudicial requerida':
                                    ws11.cell(row=ws11_row_index, column=1, value=data_consulta)
                                    ws11.cell(row=ws11_row_index, column=3, value=cnpj_lista)
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes extrajudiciais concedidas':
                                    ws12.cell(row=ws12_row_index, column=1, value=data_consulta)
                                    ws12.cell(row=ws12_row_index, column=3, value=cnpj_lista)

                                for item in paragrafo_separado: 

                                    if ':' in item: 
                                        if 'CNPJ' not in item:
                                            item_movimento = item.split(":")
                                            titulo = item_movimento[0]
                                            valor = str(item_movimento[1]) if item_movimento[1] is not None else ''
                                            coluna = None

                                            if self.__remover_acentos_(titulo).lower() == 'empresa' or self.__remover_acentos_(titulo).lower() == 'requerido':
                                                coluna = 2
                                            elif self.__remover_acentos_(titulo).lower() == 'cnpj':
                                                coluna = 3
                                            elif self.__remover_acentos_(titulo).lower().strip() == 'endereco':
                                                coluna = 4
                                            elif self.__remover_acentos_(titulo).lower().strip() == 'administrador judicial' or self.__remover_acentos_(titulo).lower() == 'requerente':
                                                coluna = 5
                                            elif self.__remover_acentos_(titulo).lower().strip() == 'vara/comarca':
                                                coluna = 6
                                            elif self.__remover_acentos_(titulo).lower().strip() == 'observacao':
                                                coluna = 7

                                            if coluna is not None:

                                                if self.__remover_acentos_(i.titulo).lower() == 'falencias requeridas':
                                                    ws0.cell(row=ws0_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'falencias decretadas':
                                                    ws1.cell(row=ws1_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'processos de falencia extintos':
                                                    ws2.cell(row=ws2_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'reformas de sentenca de falencia':
                                                    ws3.cell(row=ws3_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao judicial deferida':
                                                    ws4.cell(row=ws4_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao extrajudicial deferida':
                                                    ws5.cell(row=ws5_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'homologacao de desistencia de recuperacao judicial':
                                                    ws6.cell(row=ws6_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'cumprimento de recuperacao judicial':
                                                    ws7.cell(row=ws7_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes judiciais concedidas':
                                                    ws8.cell(row=ws8_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao judicial requerida':
                                                    ws9.cell(row=ws9_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes judiciais indeferidas':
                                                    ws10.cell(row=ws10_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao extrajudicial requerida':
                                                    ws11.cell(row=ws11_row_index, column=coluna, value=valor)
                                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes extrajudiciais concedidas':
                                                    ws12.cell(row=ws12_row_index, column=coluna, value=valor)
                                        else:
                                            pass        
                                    else:
                                        pass

                                if self.__remover_acentos_(i.titulo).lower() == 'falencias requeridas':
                                    ws0_row_index = ws0_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'falencias decretadas':
                                    ws1_row_index = ws1_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'processos de falencia extintos':
                                    ws2_row_index = ws2_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'reformas de sentenca de falencia':
                                    ws3_row_index = ws3_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao judicial deferida':
                                    ws4_row_index = ws4_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao extrajudicial deferida':
                                    ws5_row_index = ws5_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'homologacao de desistencia de recuperacao judicial':
                                    ws6_row_index = ws6_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'cumprimento de recuperacao judicial':
                                    ws7_row_index = ws7_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes judiciais concedidas':
                                    ws8_row_index = ws8_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao judicial requerida':
                                    ws9_row_index = ws9_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes judiciais indeferidas':
                                    ws10_row_index = ws10_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacao extrajudicial requerida':
                                    ws11_row_index = ws11_row_index +1
                                elif self.__remover_acentos_(i.titulo).lower() == 'recuperacoes extrajudiciais concedidas':                                
                                    ws12_row_index = ws12_row_index +1

                    except Exception as ex:
                        print(f'Erro no processamento de uma pagina {str(ex)}')
                        sem_erros = True

                    contador = contador+1

                path = f'./dados_capturados/movimento_falimentar_{page}.xlsx'
                print(f'Salvando excel em: {path}')
                wb.save(path)

                total_processado = page_count
                print(f'Total processado: {total_processado}')

                page = page + 1
                time.sleep(15)
                print(f'Paginando novos resultados...')
                pagina.locator(self.BOTAO_PAGINACAO).click(timeout=3000)
                time.sleep(15)

            print('Realizando logout...')
            self.realizar_logoff(pagina=pagina)
            print('Logout realizado...')

            result = True
        except Exception as ex:
            print(f'Erro na execucao geral: {str(ex)}')

            print('Realizando logout...')
            self.realizar_logoff(pagina=pagina)
            print('Logout realizado...')

        finally:
            playwright.stop()            

        return result

    def realizar_login(self, pagina:Page) -> bool:
        result = False

        self.go_to_pagina(pagina, self.url_base)

        time.sleep(10)

        pagina.wait_for_selector(self.BOTAO_LOGIN)

        try:
            pagina.locator(self.BOTAO_LOGIN).click(timeout=5000)
        except PlaywrightTimeoutError as e:
            pass

        login:str=""
        senha:str=""

        try:        
            pagina.wait_for_selector(self.INPUT_LOGIN_SENHA)

            print('Preenchendo usuario...')
            pagina.locator(self.INPUT_LOGIN_SENHA).type(login, delay=146)
            time.sleep(2)

            try:
                pagina.locator(self.BOTAO_CONTINUAR_ENTRAR).click(timeout=5000)
            except PlaywrightTimeoutError as e:
                pass
            
            time.sleep(15)

            print('Preenchendo senha...')
            pagina.locator(self.INPUT_LOGIN_SENHA).type(senha, delay=146)
            time.sleep(2)

            try:
                pagina.locator(self.BOTAO_CONTINUAR_ENTRAR).click(timeout=5000)
            except PlaywrightTimeoutError as e:
                pass
            
            time.sleep(15)

            result = True
        except PlaywrightTimeoutError as e:
            print("Nao foi possivel realizar o login. Erro: " + str(e))
            pass

        return result

    def realizar_logoff(self, pagina:Page) -> bool:
        result = False

        try:
            self.go_to_pagina(pagina, self.url_base)
                
            time.sleep(3)

            menu_sair = pagina.locator('.login-profile__name')

            menu_sair.hover()

            pagina.locator(self.BOTAO_SAIR).nth(3).click()

            result = True
        except PlaywrightTimeoutError as e:
            print("Nao foi possivel realizar o logout")
            pass

        return result

    def __obter_workbook() -> Workbook:
        return Workbook()

        # ws0 = wb.create_sheet("Falencia Requerida")
        # ws0.cell(row=1, column=1, value='DATA')
        # ws0.cell(row=1, column=2, value='REQUERIDO')
        # ws0.cell(row=1, column=3, value='CNPJ')
        # ws0.cell(row=1, column=4, value='ENDERECO')
        # ws0.cell(row=1, column=5, value='REQUERENTE')
        # ws0.cell(row=1, column=6, value='VARA/COMARCA')
        # ws0.cell(row=1, column=7, value='OBS')
        # ws0_row_index = 2

        # ws1 = wb.create_sheet("Falencia Decretada")
        # ws1.cell(row=1, column=1, value='DATA')
        # ws1.cell(row=1, column=2, value='EMPRESA')
        # ws1.cell(row=1, column=3, value='CNPJ')
        # ws1.cell(row=1, column=4, value='ENDERECO')
        # ws1.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws1.cell(row=1, column=6, value='VARA/COMARCA')
        # ws1.cell(row=1, column=7, value='OBS')
        # ws1_row_index = 2

        # ws2 = wb.create_sheet("Processos Falencia Extintos")
        # ws2.cell(row=1, column=1, value='DATA')
        # ws2.cell(row=1, column=2, value='REQUERIDO')
        # ws2.cell(row=1, column=3, value='CNPJ')
        # ws2.cell(row=1, column=4, value='ENDERECO')
        # ws2.cell(row=1, column=5, value='REQUERENTE')
        # ws2.cell(row=1, column=6, value='VARA/COMARCA')
        # ws2.cell(row=1, column=7, value='OBS')
        # ws2_row_index = 2
                 
        # ws3 = wb.create_sheet("Reformas de Sentenca de Falencia")
        # ws3.cell(row=1, column=1, value='DATA')
        # ws3.cell(row=1, column=2, value='REQUERIDO')
        # ws3.cell(row=1, column=3, value='CNPJ')
        # ws3.cell(row=1, column=4, value='ENDERECO')
        # ws3.cell(row=1, column=5, value='REQUERENTE')
        # ws3.cell(row=1, column=6, value='VARA/COMARCA')
        # ws3.cell(row=1, column=7, value='OBS')
        # ws3_row_index = 2

        # ws4 = wb.create_sheet("RJ Deferida")
        # ws4.cell(row=1, column=1, value='DATA')
        # ws4.cell(row=1, column=2, value='EMPRESA')
        # ws4.cell(row=1, column=3, value='CNPJ')
        # ws4.cell(row=1, column=4, value='ENDERECO')
        # ws4.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws4.cell(row=1, column=6, value='VARA/COMARCA')
        # ws4.cell(row=1, column=7, value='OBS')
        # ws4_row_index = 2

        # ws5 = wb.create_sheet("RE Deferida")
        # ws5.cell(row=1, column=1, value='DATA')
        # ws5.cell(row=1, column=2, value='EMPRESA')
        # ws5.cell(row=1, column=3, value='CNPJ')
        # ws5.cell(row=1, column=4, value='ENDERECO')
        # ws5.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws5.cell(row=1, column=6, value='VARA/COMARCA')
        # ws5.cell(row=1, column=7, value='OBS')
        # ws5_row_index = 2

        # ws6 = wb.create_sheet("Homologacao de Desistencia RJ")
        # ws6.cell(row=1, column=1, value='DATA')
        # ws6.cell(row=1, column=2, value='EMPRESA')
        # ws6.cell(row=1, column=3, value='CNPJ')
        # ws6.cell(row=1, column=4, value='ENDERECO')
        # ws6.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws6.cell(row=1, column=6, value='VARA/COMARCA')
        # ws6.cell(row=1, column=7, value='OBS')
        # ws6_row_index = 2

        # ws7 = wb.create_sheet("Cumprimento de RJ")
        # ws7.cell(row=1, column=1, value='DATA')
        # ws7.cell(row=1, column=2, value='EMPRESA')
        # ws7.cell(row=1, column=3, value='CNPJ')
        # ws7.cell(row=1, column=4, value='ENDERECO')
        # ws7.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws7.cell(row=1, column=6, value='VARA/COMARCA')
        # ws7.cell(row=1, column=7, value='OBS')
        # ws7_row_index = 2

        # ws8 = wb.create_sheet("RJ Concedidas")
        # ws8.cell(row=1, column=1, value='DATA')
        # ws8.cell(row=1, column=2, value='EMPRESA')
        # ws8.cell(row=1, column=3, value='CNPJ')
        # ws8.cell(row=1, column=4, value='ENDERECO')
        # ws8.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws8.cell(row=1, column=6, value='VARA/COMARCA')
        # ws8.cell(row=1, column=7, value='OBS')
        # ws8_row_index = 2

        # ws9 = wb.create_sheet("RJ Requerida")
        # ws9.cell(row=1, column=1, value='DATA')
        # ws9.cell(row=1, column=2, value='EMPRESA')
        # ws9.cell(row=1, column=3, value='CNPJ')
        # ws9.cell(row=1, column=4, value='ENDERECO')
        # ws9.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws9.cell(row=1, column=6, value='VARA/COMARCA')
        # ws9.cell(row=1, column=7, value='OBS')
        # ws9_row_index = 2

        # ws10 = wb.create_sheet("RJ Indeferida")
        # ws10.cell(row=1, column=1, value='DATA')
        # ws10.cell(row=1, column=2, value='EMPRESA')
        # ws10.cell(row=1, column=3, value='CNPJ')
        # ws10.cell(row=1, column=4, value='ENDERECO')
        # ws10.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws10.cell(row=1, column=6, value='VARA/COMARCA')
        # ws10.cell(row=1, column=7, value='OBS')
        # ws10_row_index = 2

        # ws11 = wb.create_sheet("RE Requerida")
        # ws11.cell(row=1, column=1, value='DATA')
        # ws11.cell(row=1, column=2, value='EMPRESA')
        # ws11.cell(row=1, column=3, value='CNPJ')
        # ws11.cell(row=1, column=4, value='ENDERECO')
        # ws11.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws11.cell(row=1, column=6, value='VARA/COMARCA')
        # ws11.cell(row=1, column=7, value='OBS')
        # ws11_row_index = 2

        # ws12 = wb.create_sheet("RE Concedidas")
        # ws12.cell(row=1, column=1, value='DATA')
        # ws12.cell(row=1, column=2, value='EMPRESA')
        # ws12.cell(row=1, column=3, value='CNPJ')
        # ws12.cell(row=1, column=4, value='ENDERECO')
        # ws12.cell(row=1, column=5, value='ADM JUDICIAL')
        # ws12.cell(row=1, column=6, value='VARA/COMARCA')
        # ws12.cell(row=1, column=7, value='OBS')
        # ws12_row_index = 2

    def __create_sheet(self, wb: Workbook, titulo: str):

        ws = wb.create_sheet(titulo)
        ws.cell(row=1, column=1, value='DATA')
        ws.cell(row=1, column=2, value='REQUERIDO')
        ws.cell(row=1, column=3, value='CNPJ')
        ws.cell(row=1, column=4, value='ENDERECO')
        ws.cell(row=1, column=5, value='REQUERENTE')
        ws.cell(row=1, column=6, value='VARA/COMARCA')
        ws.cell(row=1, column=7, value='OBS')
        ws_row_index = 2

        return ws, ws_row_index

    def go_to_pagina(self, pagina:Page, url:str, timeout:int = 5000):
        if (pagina.url != url):
            try:
                pagina.goto(url=url, wait_until='domcontentloaded', timeout=timeout)
            except PlaywrightTimeoutError as e:
                pass

            time.sleep(2)

    def __remover_acentos_(self, source:str) -> str:
        result = ''

        if source != None:
            result = normalize('NFKD', source).encode('ASCII','ignore').decode('ASCII')

        return result

